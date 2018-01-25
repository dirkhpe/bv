"""
This script will analyze the solution to solution information for loading in Neo4J.
"""

import logging
import pymsgbox as msg
from lib import my_env
from lib import localstore
from lib.localstore import *
from openpyxl import load_workbook
# import sqlalchemy
from sqlalchemy.exc import OperationalError


msg_title = "Warning!"
msg_text = "Convert MURCS xls to xlsx. Do not use MURCS generated xlsx."
msg.alert(title=msg_title, text=msg_text, button="OK")
cfg = my_env.init_env("bellavista", __file__)
logging.info("Start Application")
db = cfg['Main']['db']
ds, engine = localstore.init_session(db=db)
try:
    SolutionToSolution.__table__.drop(bind=engine)
except OperationalError:
    pass
SolutionToSolution.__table__.create(bind=engine)
try:
    Solution.__table__.drop(bind=engine)
except:
    pass
Solution.__table__.create(bind=engine)
# First handle solutions
wb = load_workbook(cfg['Main']['solutions'], read_only=True)
sheet = wb['Sheet1']
# Get worksheet data as a list of rows
row_list = list(sheet.rows)
# Get title row in a named row (nr) to handle all data rows as dictionaries
title_row = row_list.pop(0)
nr = my_env.get_named_row("solutions", title_row)
# Then handle each record in the sheet
loop = my_env.LoopInfo("solutions", 25)
for row in map(nr._make, row_list):
    loop.info_loop()
    sol = Solution(
        solId=row.solId.value,
        name=row.solName.value,
        comment=row.comment.value,
        complexity=row.complexity.value,
        inScope=row.inScope.value,
        origin=row.origin.value,
        customerBusinessUnit=row.customerBusinessUnit.value,
        customerBusinessDivision=row.customerBusinessDivision.value,
        supportBusinessUnit=row.supportBusinessUnit.value,
        supportBusinessDivision=row.supportBusinessDivision.value,
        longDescription=row.longDescription.value,
        description=row.description.value,
        applicationTreatment=row.applicationTreatment.value,
        classification=row.classification.value
    )
    ds.add(sol)
ds.commit()
loop.end_loop()

# Then handle solution to solution
wb = load_workbook(cfg['Main']['sol2sol'], read_only=True)
sheet = wb['Sheet1']
# Get worksheet data as a list of rows
row_list = list(sheet.rows)
# Get title row in a named row (nr) to handle all data rows as dictionaries
title_row = row_list.pop(0)
nr = my_env.get_named_row("sol2sol", title_row)
# Then handle each record in the sheet
loop = my_env.LoopInfo("sol2sol", 25)
for row in map(nr._make, row_list):
    loop.info_loop()
    sol2sol = SolutionToSolution(
        fromSolId=row.fromSolId.value,
        fromSolName=row.fromSolName.value,
        toSolId=row.toSolId.value,
        toSolName=row.toSolName.value,
        conType=row.conType.value,
        conSubType=row.conSubType.value,
        conDirection=row.conDirection.value,
        middlewareDependency=row.middlewareDependency.value,
        comment=row.comment.value,
        description=row.description.value
    )
    ds.add(sol2sol)
ds.commit()
loop.end_loop()

